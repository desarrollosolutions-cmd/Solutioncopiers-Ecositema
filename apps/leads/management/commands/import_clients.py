"""
Importa clientes desde las 3 bases de datos Excel y los fusiona en un solo Lead por NIT.
Los campos sin datos quedan vacíos para completar manualmente.

Archivos esperados (rutas por defecto o con --dir):
  - LIstado de terceros siigo*.xlsx
  - listado de terceros worffice*.xlsx
  - listado de terceros x mariana*.xlsx

Uso:
    python manage.py import_clients
    python manage.py import_clients --dir "C:/ruta/a/carpeta"
    python manage.py import_clients --dry-run
"""
from __future__ import annotations

import glob
import os
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except ImportError:
    raise CommandError("Instala openpyxl: pip install openpyxl")

# ---------------------------------------------------------------------------
# Mapeo de códigos DANE → nombre de ciudad (principales ciudades)
# ---------------------------------------------------------------------------
_DANE = {
    "05001": "Medellín", "05088": "Bello", "05380": "La Estrella",
    "05266": "Envigado", "05360": "Itagüí", "05631": "Sabaneta",
    "05129": "Caldas", "05212": "Copacabana", "05615": "Rionegro",
    "05045": "Apartadó", "05790": "Turbo", "05001001": "Medellín",
    "11001": "Bogotá", "76001": "Cali", "08001": "Barranquilla",
    "13001": "Cartagena", "68001": "Bucaramanga", "17001": "Manizales",
    "63001": "Armenia", "66001": "Pereira", "52001": "Pasto",
    "73001": "Ibagué", "54001": "Cúcuta", "23001": "Montería",
    "70001": "Sincelejo", "41001": "Neiva", "15001": "Tunja",
    "18001": "Florencia", "85001": "Yopal", "86001": "Mocoa",
}

DEFAULT_DIR = r"C:\Users\User\Desktop\base de datos solutions"


def _clean_nit(raw: str) -> str:
    """'NIT 900746054 4' → '900746054'  |  '860045904-7' → '860045904'"""
    if not raw:
        return ""
    s = str(raw).strip()
    s = re.sub(r"(?i)nit\s*", "", s)
    s = re.sub(r"(?i)c[eé]dula\s*", "", s)
    s = s.replace("-", " ").strip()
    parts = s.split()
    return parts[0] if parts else ""


def _clean_phone(raw) -> str:
    if not raw:
        return ""
    phone = re.sub(r"[^\d+]", "", str(raw))
    if phone.startswith("57") and len(phone) > 10:
        phone = phone[2:]
    return phone[:20]


def _city_name(code) -> str:
    if not code:
        return ""
    return _DANE.get(str(code).strip(), "")


# ---------------------------------------------------------------------------
# Lectores por archivo
# ---------------------------------------------------------------------------

def _read_siigo(path: str) -> list[dict]:
    """Lee el export de Siigo, filtra solo Clientes=SI."""
    records = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if len(row) < 31:
            continue
        # col 29 (index 28) = Clientes
        if str(row[28]).strip().upper() != "SI":
            continue

        tipo       = str(row[4] or "").strip()
        razon      = str(row[5] or "").strip()
        nombres    = str(row[6] or "").strip()
        apellidos  = str(row[7] or "").strip()
        nombre_com = str(row[8] or "").strip()
        direccion  = str(row[9] or "").strip()
        ciudad_cod = str(row[12] or "").strip()
        telefono   = _clean_phone(row[14])
        nom_cont   = str(row[19] or "").strip()
        ape_cont   = str(row[20] or "").strip()
        tel_cont   = _clean_phone(row[22])
        email      = str(row[24] or "").strip()
        nit        = _clean_nit(str(row[0] or ""))
        estado     = str(row[30] or "").strip().lower()

        if estado == "inactivo":
            continue

        # Nombre completo: contacto principal > razón social
        if nom_cont:
            full_name = f"{nom_cont} {ape_cont}".strip()
            company   = razon or nombre_com
        elif nombres:
            full_name = f"{nombres} {apellidos}".strip()
            company   = razon or nombre_com
        else:
            full_name = razon or nombre_com
            company   = razon

        phone_final = telefono or tel_cont
        city_name   = _city_name(ciudad_cod)

        notes = ""
        if direccion:
            notes = f"Dirección: {direccion}"

        records.append({
            "nit":          nit,
            "full_name":    full_name[:150],
            "company_name": company[:200],
            "email":        email[:254] if "@" in email else "",
            "phone":        phone_final,
            "city":         city_name or "Medellín",
            "notes_internal": notes,
            "_source_file": "siigo",
        })
    wb.close()
    return records


def _read_worffice_style(path: str, source_tag: str) -> list[dict]:
    """Lee exports de Worffice/Mariana: cols Clasificacion1,2, Nombre, Identificacion, Propiedades."""
    records = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_found = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        # Buscar la fila de encabezado
        if not header_found:
            if row[0] and "clasificacion" in str(row[0]).lower():
                header_found = True
            continue

        if not row[2]:  # sin nombre
            continue

        props = str(row[4] or "").lower()
        if "cliente" not in props:
            continue

        nombre = str(row[2] or "").strip()
        nit    = _clean_nit(str(row[3] or ""))

        records.append({
            "nit":          nit,
            "full_name":    nombre[:150],
            "company_name": nombre[:200],
            "email":        "",
            "phone":        "",
            "city":         "Medellín",
            "notes_internal": "",
            "_source_file": source_tag,
        })
    wb.close()
    return records


# ---------------------------------------------------------------------------
# Fusión y deduplicación
# ---------------------------------------------------------------------------

def _merge_records(all_records: list[dict]) -> list[dict]:
    """
    Deduplica por NIT. Siigo tiene prioridad (datos más ricos).
    Sin NIT: se importan todos (no hay forma de detectar duplicados).
    """
    seen_nit: dict[str, dict] = {}
    no_nit: list[dict] = []

    for rec in all_records:
        nit = rec["nit"]
        if not nit:
            no_nit.append(rec)
            continue
        if nit in seen_nit:
            existing = seen_nit[nit]
            # Completar campos vacíos con los del nuevo registro
            for field in ("email", "phone", "city", "notes_internal"):
                if not existing[field] and rec[field]:
                    existing[field] = rec[field]
        else:
            seen_nit[nit] = rec

    return list(seen_nit.values()) + no_nit


# ---------------------------------------------------------------------------
# Command
# ---------------------------------------------------------------------------

class Command(BaseCommand):
    help = "Importa clientes desde los 3 archivos Excel y los fusiona en una sola base"

    def add_arguments(self, parser):
        parser.add_argument(
            "--dir",
            default=DEFAULT_DIR,
            help="Carpeta donde están los archivos Excel (default: escritorio)",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Solo muestra estadísticas, no guarda nada",
        )

    def handle(self, *args, **options):
        base_dir = options["dir"]
        dry_run  = options["dry_run"]

        if not os.path.isdir(base_dir):
            raise CommandError(f"Carpeta no encontrada: {base_dir}")

        # ── Localizar archivos ──────────────────────────────────────────
        def find(pattern):
            matches = glob.glob(os.path.join(base_dir, pattern), recursive=False)
            return matches[0] if matches else None

        siigo_path    = find("*siigo*")
        worffice_path = find("*worffice*")
        mariana_path  = find("*mariana*")

        if not siigo_path:
            raise CommandError("No se encontró el archivo de Siigo en la carpeta indicada.")

        # ── Leer archivos ───────────────────────────────────────────────
        self.stdout.write("Leyendo Siigo…")
        siigo_recs = _read_siigo(siigo_path)
        self.stdout.write(f"  -> {len(siigo_recs)} clientes activos en Siigo")

        worffice_recs = []
        if worffice_path:
            self.stdout.write("Leyendo Worffice…")
            worffice_recs = _read_worffice_style(worffice_path, "worffice")
            self.stdout.write(f"  -> {len(worffice_recs)} clientes en Worffice")

        mariana_recs = []
        if mariana_path:
            self.stdout.write("Leyendo Mariana…")
            mariana_recs = _read_worffice_style(mariana_path, "mariana")
            self.stdout.write(f"  -> {len(mariana_recs)} clientes en Mariana")

        # Siigo primero (datos más ricos)
        all_records = siigo_recs + worffice_recs + mariana_recs
        merged = _merge_records(all_records)
        self.stdout.write(f"\nRegistros únicos después de deduplicar por NIT: {len(merged)}")

        if dry_run:
            self.stdout.write(self.style.WARNING("--dry-run: no se guardaron datos."))
            return

        # ── Importar a la base de datos ─────────────────────────────────
        from apps.leads.models import Lead

        created = updated = skipped = 0

        for rec in merged:
            nit        = rec["nit"]
            full_name  = rec["full_name"]
            email      = rec["email"]

            if not full_name:
                skipped += 1
                continue

            # Intentar encontrar existente por NIT (primero), luego por email
            existing = None
            if nit:
                existing = Lead.objects.filter(nit=nit).first()
            if not existing and email:
                existing = Lead.objects.filter(email=email).first()

            if existing:
                # Completar campos vacíos sin pisar datos existentes
                changed = False
                for field, val in [
                    ("phone",        rec["phone"]),
                    ("email",        rec["email"]),
                    ("company_name", rec["company_name"]),
                    ("city",         rec["city"]),
                    ("nit",          nit),
                ]:
                    if val and not getattr(existing, field):
                        setattr(existing, field, val)
                        changed = True
                if changed:
                    existing.save()
                    updated += 1
                else:
                    skipped += 1
            else:
                Lead.objects.create(
                    full_name    = full_name,
                    email        = email,
                    phone        = rec["phone"],
                    company_name = rec["company_name"],
                    city         = rec["city"],
                    nit          = nit,
                    notes_internal = rec["notes_internal"],
                    source       = Lead.Source.IMPORT,
                    gdpr_consent = False,
                )
                created += 1

        self.stdout.write(self.style.SUCCESS(
            f"\nImportación completada:\n"
            f"  Creados:       {created}\n"
            f"  Actualizados:  {updated}\n"
            f"  Sin cambios:   {skipped}\n"
            f"  Total procesados: {created + updated + skipped}"
        ))
