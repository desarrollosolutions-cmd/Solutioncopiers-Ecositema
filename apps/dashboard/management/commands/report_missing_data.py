from django.core.management.base import BaseCommand


class Command(BaseCommand):
    help = "Exporta a un único Excel (2 hojas) consumibles sin precio y clientes sin correo"

    def add_arguments(self, parser):
        parser.add_argument("--outdir", default=".", help="Directorio de salida")
        parser.add_argument("--filename", default="datos_faltantes.xlsx", help="Nombre del archivo Excel")

    def handle(self, *args, **options):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from apps.catalog.models import Consumable
        from apps.leads.models import Lead

        outdir   = options["outdir"]
        filename = options["filename"]
        path     = f"{outdir}/{filename}"

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        def style_header(ws, ncols):
            for col in range(1, ncols + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font

        wb = Workbook()

        # --- Hoja 1: Consumibles sin precio ---
        ws1 = wb.active
        ws1.title = "Consumibles sin precio"
        ws1.append(["Nombre", "Referencia"])
        cons_no_price = Consumable.objects.filter(price__isnull=True).order_by("name")
        for c in cons_no_price:
            ws1.append([c.name, c.part_number])
        style_header(ws1, 2)
        ws1.column_dimensions["A"].width = 70
        ws1.column_dimensions["B"].width = 35
        ws1.freeze_panes = "A2"

        # --- Hoja 2: Clientes sin correo ---
        ws2 = wb.create_sheet("Clientes sin correo")
        ws2.append(["Nombre", "Empresa", "Teléfono", "NIT"])
        leads_no_email = Lead.objects.filter(email="").order_by("full_name")
        total_leads = Lead.objects.count()
        for l in leads_no_email:
            ws2.append([l.full_name, l.company_name, l.phone, l.nit])
        style_header(ws2, 4)
        ws2.column_dimensions["A"].width = 45
        ws2.column_dimensions["B"].width = 45
        ws2.column_dimensions["C"].width = 18
        ws2.column_dimensions["D"].width = 18
        ws2.freeze_panes = "A2"

        wb.save(path)

        self.stdout.write(self.style.SUCCESS(
            f"Consumibles sin precio: {cons_no_price.count()}"
        ))
        self.stdout.write(self.style.SUCCESS(
            f"Clientes sin correo: {leads_no_email.count()} de {total_leads} totales"
        ))
        self.stdout.write(self.style.SUCCESS(f"Archivo generado: {path}"))
